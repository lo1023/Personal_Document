import os
import openpyxl as xl

#__file__ 현재 수행중인 파이썬의 경로를 반환한다.
# 파이썬은 {}가 없는대신 들여쓰기 사용한다.
# if 나 for 쓰고 난 후에 들여쓰기 한칸 무조건

currentPath = os.path.dirname(__file__)
reportPath = os.path.join(currentPath,'data')


reports = [] #가변 배열 쩌..쩐당

for file in os.listdir(reportPath):
    if file.endswith(".xlsx") and "20200104" in file:
        filePath = os.path.join(reportPath,file)

        wb = xl.load_workbook(filePath)
        sheet = wb.active

        #3행부터 1열 - 5열
        row = 3
        col = 1

        #10번 반복
        while 1:
            name = sheet.cell(row = row, column = col).value
            if name is None:
                break
            content = sheet.cell(row = row, column = col+1).value
            reports.append({"name" : name,"content" : content}) # 이렇게 하면 자동으로 구조체가 형성되는건가? append가 단순 문자열 합치기는 아닌가?
            row +=1


NameList = {} # [] {}
newRow = 0

wb = xl.load_workbook(os.path.join(reportPath,"20200103.xlsx"))
sheet = wb.active

row = 3
while 1:
    name = sheet.cell(row = row, column = col).value
    if name is None:
        newRow = row
        break

    NameList[name] = row  #[name] == C++ Temp[row] = name 인
    row += 1


col = 1
for r in reports:
    row = NameList.get(r["name"] , -1) # tame[n] == "name" 있으면, 인덱스 반환 없으면 -1
    if row == -1:
        row = newRow
        newRow += 1


    sheet.cell(row = row , column = col).value = r["name"]
    sheet.cell(row = row , column = col+1).value = r["content"]

wb.save(os.path.join(reportPath,"20200104.xlsx"))


"""
1. 파이썬에서 append의 의미는 무엇인가?
"""